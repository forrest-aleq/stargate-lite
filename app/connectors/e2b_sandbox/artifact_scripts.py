# ruff: noqa: E501
"""Pure-Python artifact builders that run inside an E2B sandbox."""

from __future__ import annotations

import base64
import json
from textwrap import dedent
from typing import Any


def _encode_spec(spec: dict[str, Any]) -> str:
    payload = json.dumps(spec, separators=(",", ":"), ensure_ascii=True).encode("utf-8")
    return base64.b64encode(payload).decode("ascii")


def build_xlsx_script(spec: dict[str, Any]) -> str:
    encoded = _encode_spec(spec)
    template = """
    import base64
    import io
    import json
    import re
    import subprocess
    import sys
    from pathlib import Path

    SPEC = json.loads(base64.b64decode("__ENCODED__").decode("utf-8"))


    def ensure_openpyxl():
        try:
            from openpyxl import Workbook
            from openpyxl.chart import BarChart, LineChart
            from openpyxl.chart.reference import Reference
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter, range_boundaries
            from openpyxl.workbook.defined_name import DefinedName
            from openpyxl.workbook.properties import CalcProperties
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except ImportError:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", "openpyxl==3.1.5"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
            from openpyxl import Workbook
            from openpyxl.chart import BarChart, LineChart
            from openpyxl.chart.reference import Reference
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter, range_boundaries
            from openpyxl.workbook.defined_name import DefinedName
            from openpyxl.workbook.properties import CalcProperties
            from openpyxl.worksheet.table import Table, TableStyleInfo
        return (
            Workbook,
            BarChart,
            LineChart,
            Reference,
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
            get_column_letter,
            range_boundaries,
            DefinedName,
            CalcProperties,
            Table,
            TableStyleInfo,
        )


    (
        Workbook,
        BarChart,
        LineChart,
        Reference,
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
        get_column_letter,
        range_boundaries,
        DefinedName,
        CalcProperties,
        Table,
        TableStyleInfo,
    ) = ensure_openpyxl()


    HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF3F51B5")
    HEADER_FONT = Font(bold=True, color="FFFFFFFF", name="Aptos", size=11)
    HEADER_BORDER = Border(
        top=Side(style="thin", color="FFD0D7EE"),
        bottom=Side(style="thin", color="FFD0D7EE"),
    )
    HEADER_ALIGNMENT = Alignment(vertical="center")
    NUMBER_FORMATS = {
        "currency": "$#,##0.00;[Red]-$#,##0.00",
        "percent": "0.0%",
        "integer": "#,##0",
    }
    CELL_SPEC_KEYS = {"value", "formula", "style", "number_format", "hyperlink"}


    def normalize_columns(rows, explicit):
        columns = []
        if explicit:
            for index, raw_column in enumerate(explicit, start=1):
                if isinstance(raw_column, dict) and not CELL_SPEC_KEYS.intersection(raw_column.keys()):
                    header = str(
                        raw_column.get("header")
                        or raw_column.get("name")
                        or raw_column.get("label")
                        or f"Column {index}"
                    )
                    columns.append(
                        {
                            "header": header,
                            "width": raw_column.get("width"),
                        }
                    )
                else:
                    columns.append({"header": str(raw_column), "width": None})
            return columns

        seen = set()
        for row in rows:
            if isinstance(row, dict) and not CELL_SPEC_KEYS.intersection(row.keys()):
                for key in row:
                    key_text = str(key)
                    if key_text not in seen:
                        seen.add(key_text)
                        columns.append({"header": key_text, "width": None})
        return columns


    def normalize_rows(rows, columns):
        normalized = []
        for row in rows:
            if isinstance(row, dict) and not CELL_SPEC_KEYS.intersection(row.keys()):
                normalized.append([row.get(column) for column in columns])
            elif isinstance(row, list):
                normalized.append(row[: len(columns)] + [None] * max(0, len(columns) - len(row)))
            else:
                normalized.append([row] + [None] * max(0, len(columns) - 1))
        return normalized


    def bool_spec(value, default):
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            lowered = value.strip().lower()
            if lowered in {"true", "1", "yes", "on"}:
                return True
            if lowered in {"false", "0", "no", "off"}:
                return False
        raise ValueError("Boolean setting is malformed")


    def cell_spec(value):
        return value if isinstance(value, dict) and CELL_SPEC_KEYS.intersection(value.keys()) else None


    def normalize_formula(value):
        text = str(value).strip()
        if not text:
            raise ValueError("Formula cannot be empty")
        return text if text.startswith("=") else "=" + text


    def apply_style(cell, style_name, explicit_number_format=None):
        if explicit_number_format:
            cell.number_format = str(explicit_number_format)
            return
        if style_name in NUMBER_FORMATS:
            cell.number_format = NUMBER_FORMATS[style_name]


    def set_cell(cell, raw_value, inferred_style):
        spec = cell_spec(raw_value)
        if spec is None:
            cell.value = raw_value
            apply_style(cell, inferred_style)
            return bool(isinstance(raw_value, str) and raw_value.startswith("="))

        style_name = str(spec.get("style") or inferred_style or "").strip().lower() or None
        explicit_number_format = spec.get("number_format")
        if "formula" in spec and spec.get("formula") is not None:
            cell.value = normalize_formula(spec["formula"])
            apply_style(cell, style_name, explicit_number_format)
            if spec.get("hyperlink"):
                cell.hyperlink = str(spec["hyperlink"])
            return True

        cell.value = spec.get("value")
        apply_style(cell, style_name, explicit_number_format)
        if spec.get("hyperlink"):
            cell.hyperlink = str(spec["hyperlink"])
        return False


    def unique_sheet_title(raw_name, existing):
        base = str(raw_name or "Sheet").strip() or "Sheet"
        base = re.sub(r"[\\\\/*?:\\[\\]]", "-", base)[:31] or "Sheet"
        candidate = base
        index = 2
        while candidate in existing:
            suffix = f" {index}"
            candidate = f"{base[: 31 - len(suffix)]}{suffix}"
            index += 1
        existing.add(candidate)
        return candidate


    def unique_table_name(raw_name, existing):
        base = re.sub(r"[^A-Za-z0-9_]", "_", str(raw_name or "AleqTable")).strip("_") or "AleqTable"
        if base[0].isdigit():
            base = f"T_{base}"
        candidate = base[:255]
        index = 2
        while candidate in existing:
            suffix = f"_{index}"
            candidate = f"{base[: 255 - len(suffix)]}{suffix}"
            index += 1
        existing.add(candidate)
        return candidate


    def resolve_range(workbook, current_sheet, raw_reference):
        text = str(raw_reference or "").strip()
        if not text:
            raise ValueError("Range reference is required")
        if "!" in text:
            sheet_name, coord = text.split("!", 1)
            sheet_name = sheet_name.strip().strip("'")
            target_sheet = workbook[sheet_name]
        else:
            target_sheet = current_sheet
            coord = text
        min_col, min_row, max_col, max_row = range_boundaries(coord)
        return Reference(
            target_sheet,
            min_col=min_col,
            min_row=min_row,
            max_col=max_col,
            max_row=max_row,
        )


    def add_charts(workbook, worksheet, chart_specs):
        if not isinstance(chart_specs, list):
            return 0
        created = 0
        for chart_spec in chart_specs:
            if not isinstance(chart_spec, dict):
                continue
            chart_type = str(chart_spec.get("type") or "line").strip().lower()
            if chart_type == "column":
                chart = BarChart()
                chart.type = "col"
            elif chart_type == "bar":
                chart = BarChart()
                chart.type = "bar"
            else:
                chart = LineChart()

            chart.title = str(chart_spec.get("title") or "")
            chart.y_axis.title = str(chart_spec.get("y_axis_title") or "")
            chart.x_axis.title = str(chart_spec.get("x_axis_title") or "")
            anchor = str(chart_spec.get("anchor") or "F2")

            series_specs = chart_spec.get("series")
            if isinstance(series_specs, list) and series_specs:
                for series_spec in series_specs:
                    if not isinstance(series_spec, dict):
                        continue
                    values = resolve_range(workbook, worksheet, series_spec.get("values"))
                    chart.add_data(values, titles_from_data=bool_spec(series_spec.get("titles_from_data"), True))
                    if series_spec.get("categories"):
                        chart.set_categories(resolve_range(workbook, worksheet, series_spec.get("categories")))
            elif chart_spec.get("data_range"):
                values = resolve_range(workbook, worksheet, chart_spec.get("data_range"))
                chart.add_data(values, titles_from_data=bool_spec(chart_spec.get("titles_from_data"), True))
                if chart_spec.get("category_range"):
                    chart.set_categories(resolve_range(workbook, worksheet, chart_spec.get("category_range")))
            else:
                continue

            worksheet.add_chart(chart, anchor)
            created += 1
        return created


    sheets = SPEC.get("sheets") or []
    if not sheets:
        raise ValueError("At least one sheet is required")

    workbook_name = str(SPEC.get("workbook_name") or "aleq-workbook.xlsx")
    if not workbook_name.lower().endswith(".xlsx"):
        workbook_name += ".xlsx"
    path = str(SPEC.get("path") or f"/workspace/{workbook_name}")
    Path(path).parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    workbook.properties.creator = str(SPEC.get("author") or "Aleq")
    workbook.properties.title = workbook_name

    existing_sheet_names = set()
    existing_table_names = set()
    formula_count = 0
    table_count = 0
    chart_count = 0
    sheet_names = []
    sheet_summaries = []

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_index, sheet_spec in enumerate(sheets):
        if not isinstance(sheet_spec, dict):
            raise ValueError("Each sheet must be an object")
        rows = list(sheet_spec.get("rows") or [])
        column_specs = normalize_columns(rows, sheet_spec.get("columns") or [])
        if not column_specs:
            column_specs = [{"header": "Value", "width": None}]
        columns = [str(col.get("header") or "Value") for col in column_specs]

        sheet_name = unique_sheet_title(sheet_spec.get("name") or f"Sheet {sheet_index + 1}", existing_sheet_names)
        worksheet = workbook.create_sheet(title=sheet_name)
        sheet_names.append(sheet_name)

        currency_columns = {str(col).lower() for col in sheet_spec.get("currency_columns", [])}
        percent_columns = {str(col).lower() for col in sheet_spec.get("percent_columns", [])}
        integer_columns = {str(col).lower() for col in sheet_spec.get("integer_columns", [])}
        include_header = bool_spec(sheet_spec.get("include_header"), True)
        freeze_header = bool_spec(sheet_spec.get("freeze_header"), include_header)
        body_rows = normalize_rows(rows, columns)
        sheet_formula_count = 0
        sheet_table_count = 0
        sheet_chart_count = 0

        widths = []
        for index, column_spec in enumerate(column_specs):
            explicit_width = column_spec.get("width")
            if isinstance(explicit_width, (int, float)) and explicit_width > 0:
                widths.append(float(explicit_width))
            else:
                widths.append(len(columns[index]))

        current_row = 1
        if include_header:
            for col_index, header in enumerate(columns, start=1):
                cell = worksheet.cell(row=1, column=col_index, value=str(header))
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = HEADER_BORDER
                cell.alignment = HEADER_ALIGNMENT
            current_row = 2

        for row_index, row in enumerate(body_rows, start=current_row):
            for col_index, header in enumerate(columns, start=1):
                value = row[col_index - 1] if col_index - 1 < len(row) else None
                header_key = str(header).lower()
                inferred_style = None
                if header_key in currency_columns:
                    inferred_style = "currency"
                elif header_key in percent_columns:
                    inferred_style = "percent"
                elif header_key in integer_columns:
                    inferred_style = "integer"
                if set_cell(worksheet.cell(row=row_index, column=col_index), value, inferred_style):
                    formula_count += 1
                    sheet_formula_count += 1
                spec = cell_spec(value)
                if spec is not None and "formula" in spec and spec.get("formula") is not None:
                    cell_text = str(spec["formula"])
                elif spec is not None and "value" in spec:
                    cell_text = "" if spec.get("value") is None else str(spec.get("value"))
                else:
                    cell_text = "" if value is None else str(value)
                widths[col_index - 1] = min(max(widths[col_index - 1], len(cell_text)), 42)

        for formula_spec in sheet_spec.get("formulas", []) or []:
            if not isinstance(formula_spec, dict):
                continue
            target = str(formula_spec.get("cell") or "").strip()
            if not target:
                raise ValueError("Formula cell reference is required")
            target_cell = worksheet[target]
            target_cell.value = normalize_formula(formula_spec.get("formula"))
            style_name = str(formula_spec.get("style") or "").strip().lower() or None
            apply_style(target_cell, style_name, formula_spec.get("number_format"))
            if formula_spec.get("hyperlink"):
                target_cell.hyperlink = str(formula_spec["hyperlink"])
            formula_count += 1
            sheet_formula_count += 1
            widths[target_cell.column - 1] = min(max(widths[target_cell.column - 1], len(str(formula_spec["formula"]))), 42)

        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = max(10, width + 2)

        max_row = max(worksheet.max_row, 1)
        max_col = max(worksheet.max_column, 1)
        data_ref = f"A1:{get_column_letter(max_col)}{max_row}"

        if freeze_header and include_header and worksheet.max_row >= 2:
            worksheet.freeze_panes = "A2"
        if bool_spec(sheet_spec.get("auto_filter"), include_header) and include_header and worksheet.max_row >= 1:
            worksheet.auto_filter.ref = data_ref

        table_enabled = sheet_spec.get("table_name") or bool_spec(sheet_spec.get("as_table"), False)
        if table_enabled and include_header and worksheet.max_row >= 2:
            table_name = unique_table_name(sheet_spec.get("table_name") or f"{sheet_name}_table", existing_table_names)
            table = Table(displayName=table_name, ref=data_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)
            table_count += 1
            sheet_table_count += 1

        sheet_chart_count = add_charts(workbook, worksheet, sheet_spec.get("charts"))
        chart_count += sheet_chart_count
        sheet_summaries.append(
            {
                "name": sheet_name,
                "row_count": len(body_rows),
                "column_count": len(columns),
                "formula_count": sheet_formula_count,
                "table_count": sheet_table_count,
                "chart_count": sheet_chart_count,
            }
        )

    for named_range in SPEC.get("named_ranges", []) or []:
        if not isinstance(named_range, dict):
            continue
        name = str(named_range.get("name") or "").strip()
        reference = str(named_range.get("reference") or named_range.get("formula") or "").strip()
        if not name or not reference:
            continue
        workbook.defined_names.add(DefinedName(name=name, attr_text=reference))

    buffer = io.BytesIO()
    workbook.save(buffer)
    payload = buffer.getvalue()
    Path(path).write_bytes(payload)
    print(
        json.dumps(
            {
                "file_name": workbook_name,
                "path": path,
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "size_bytes": len(payload),
                "sheet_count": len(sheet_names),
                "sheet_names": sheet_names,
                "formula_count": formula_count,
                "table_count": table_count,
                "chart_count": chart_count,
                "named_range_count": len(workbook.defined_names),
                "sheet_summaries": sheet_summaries,
                "file_content": base64.b64encode(payload).decode("ascii"),
            }
        )
    )
    """
    return dedent(template).replace("__ENCODED__", encoded).strip()


def build_chart_script(spec: dict[str, Any]) -> str:
    encoded = _encode_spec(spec)
    return dedent(
        f"""
        import base64
        import json
        import math
        from pathlib import Path
        from xml.sax.saxutils import escape

        SPEC = json.loads(base64.b64decode("{encoded}").decode("utf-8"))

        width = int(SPEC.get("width") or 960)
        height = int(SPEC.get("height") or 540)
        chart_type = str(SPEC.get("chart_type") or "line").lower()
        if chart_type == "column":
            chart_type = "bar"
        labels = [str(label) for label in SPEC.get("labels") or []]
        series = [item for item in SPEC.get("series") or [] if isinstance(item, dict)]
        if not labels:
            raise ValueError("labels are required")
        if not series:
            raise ValueError("series are required")

        file_name = str(SPEC.get("file_name") or "aleq-chart.svg")
        if not file_name.lower().endswith(".svg"):
            file_name += ".svg"
        path = str(SPEC.get("path") or f"/workspace/{{file_name}}")
        Path(path).parent.mkdir(parents=True, exist_ok=True)

        title = str(SPEC.get("title") or "Aleq chart")
        subtitle = str(SPEC.get("subtitle") or "")
        x_axis_label = str(SPEC.get("x_axis_label") or "")
        y_axis_label = str(SPEC.get("y_axis_label") or "")
        palette = ["#3F51B5", "#00A896", "#FF7A59", "#6C5CE7", "#EF476F", "#118AB2"]

        values = []
        for item in series:
            row = [float(value) for value in item.get("values") or []]
            if len(row) != len(labels):
                raise ValueError("each series must have the same number of values as labels")
            values.extend(row)
        min_val = min(0.0, min(values))
        max_val = max(0.0, max(values))
        if math.isclose(max_val, min_val):
            max_val = min_val + 1.0

        margin_left = 92
        margin_right = 36
        margin_top = 92 if subtitle else 72
        margin_bottom = 84
        plot_width = width - margin_left - margin_right
        plot_height = height - margin_top - margin_bottom

        def x_for(index):
            if len(labels) == 1:
                return margin_left + plot_width / 2
            return margin_left + (plot_width * index) / (len(labels) - 1)

        def y_for(value):
            ratio = (float(value) - min_val) / (max_val - min_val)
            return margin_top + plot_height - (ratio * plot_height)

        grid_lines = []
        y_ticks = 5
        for idx in range(y_ticks + 1):
            value = min_val + ((max_val - min_val) * idx / y_ticks)
            y = y_for(value)
            grid_lines.append(f'<line x1="{{margin_left}}" y1="{{y:.2f}}" x2="{{margin_left + plot_width}}" y2="{{y:.2f}}" stroke="#E4E8F2" stroke-width="1"/>')
            grid_lines.append(f'<text x="{{margin_left - 12}}" y="{{y + 4:.2f}}" text-anchor="end" font-size="12" fill="#6A7285">{{escape(f"{{value:,.0f}}")}}</text>')

        x_labels = []
        for idx, label in enumerate(labels):
            x = x_for(idx)
            x_labels.append(f'<text x="{{x:.2f}}" y="{{margin_top + plot_height + 24}}" text-anchor="middle" font-size="12" fill="#6A7285">{{escape(label)}}</text>')

        chart_nodes = []
        if chart_type == "line":
            for idx, item in enumerate(series):
                color = str(item.get("color") or palette[idx % len(palette)])
                points = " ".join(f"{{x_for(i):.2f}},{{y_for(v):.2f}}" for i, v in enumerate(item.get("values") or []))
                chart_nodes.append(f'<polyline fill="none" stroke="{{color}}" stroke-width="3" points="{{points}}"/>')
                for point_idx, value in enumerate(item.get("values") or []):
                    chart_nodes.append(
                        f'<circle cx="{{x_for(point_idx):.2f}}" cy="{{y_for(value):.2f}}" r="4" fill="{{color}}" stroke="white" stroke-width="2"/>'
                    )
        else:
            group_width = plot_width / max(len(labels), 1)
            bar_width = min(56, max(18, group_width / max(len(series), 1) * 0.66))
            for idx, item in enumerate(series):
                color = str(item.get("color") or palette[idx % len(palette)])
                for point_idx, value in enumerate(item.get("values") or []):
                    baseline = y_for(0)
                    top = y_for(value)
                    x = margin_left + group_width * point_idx + ((group_width - (bar_width * len(series))) / 2) + (idx * bar_width)
                    y = min(top, baseline)
                    h = abs(baseline - top)
                    chart_nodes.append(
                        f'<rect x="{{x:.2f}}" y="{{y:.2f}}" width="{{bar_width - 6:.2f}}" height="{{max(h, 1):.2f}}" rx="6" fill="{{color}}" />'
                    )

        legend = []
        legend_x = margin_left
        legend_y = 24
        for idx, item in enumerate(series):
            color = str(item.get("color") or palette[idx % len(palette)])
            name = str(item.get("name") or f"Series {{idx + 1}}")
            legend.append(f'<rect x="{{legend_x}}" y="{{legend_y}}" width="12" height="12" rx="3" fill="{{color}}"/>')
            legend.append(f'<text x="{{legend_x + 18}}" y="{{legend_y + 10}}" font-size="12" fill="#4D5566">{{escape(name)}}</text>')
            legend_x += 18 + max(56, len(name) * 7)

        svg = (
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{{width}}" height="{{height}}" viewBox="0 0 {{width}} {{height}}" role="img">'
            f'<rect width="{{width}}" height="{{height}}" rx="24" fill="#FFFFFF"/>'
            f'<text x="{{margin_left}}" y="36" font-size="28" font-weight="700" fill="#141A24">{{escape(title)}}</text>'
            + (f'<text x="{{margin_left}}" y="58" font-size="14" fill="#6A7285">{{escape(subtitle)}}</text>' if subtitle else "")
            + "".join(legend)
            + f'<line x1="{{margin_left}}" y1="{{margin_top + plot_height}}" x2="{{margin_left + plot_width}}" y2="{{margin_top + plot_height}}" stroke="#BBC4D7" stroke-width="1.5"/>'
            + f'<line x1="{{margin_left}}" y1="{{margin_top}}" x2="{{margin_left}}" y2="{{margin_top + plot_height}}" stroke="#BBC4D7" stroke-width="1.5"/>'
            + "".join(grid_lines)
            + "".join(chart_nodes)
            + "".join(x_labels)
            + (f'<text x="{{margin_left + plot_width / 2:.2f}}" y="{{height - 18}}" text-anchor="middle" font-size="12" fill="#6A7285">{{escape(x_axis_label)}}</text>' if x_axis_label else "")
            + (f'<text x="22" y="{{margin_top + plot_height / 2:.2f}}" text-anchor="middle" font-size="12" fill="#6A7285" transform="rotate(-90 22 {{margin_top + plot_height / 2:.2f}})">{{escape(y_axis_label)}}</text>' if y_axis_label else "")
            + "</svg>"
        )
        payload = svg.encode("utf-8")
        Path(path).write_bytes(payload)
        print(
            json.dumps(
                {{
                    "file_name": file_name,
                    "path": path,
                    "mime_type": "image/svg+xml",
                    "size_bytes": len(payload),
                    "chart_type": chart_type,
                    "series_count": len(series),
                    "file_content": base64.b64encode(payload).decode("ascii"),
                }}
            )
        )
        """
    ).strip()
